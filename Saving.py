from openpyxl import Workbook
def save(name, price, rating, link, i):
    # Имя файла
    filename = 'test.xlsx'
    # Cоздание Workbook
    wb = Workbook()
    sheet = wb.active
    if i == 2:
        # Вводим таблицы
        c = sheet.cell(row=1, column=1)
        c1 = sheet.cell(row=1, column=2)
        c2 = sheet.cell(row=1, column=3)
        c3 = sheet.cell(row=1, column=4)
        c4 = sheet.cell(row=1, column=5)
        # Ввод названий столбцов
        c.value = "№"
        c1.value = "Название"
        c2.value = "Цена"
        c3.value = "Рейтинг"
        c4.value = "Ссылка"
    # Занесение данных в строки
    c = sheet.cell(row=i, column=1)
    c1 = sheet.cell(row=i, column=2)
    c2 = sheet.cell(row=i, column=3)
    c3 = sheet.cell(row=i, column=4)
    c4 = sheet.cell(row=i, column=5)
    # Занесение номера
    c.value = i-1
    # Занесение имени
    c1.value = name
    # Занесение цены
    c2.value = price
    # Занесение рейтинга
    c3.value = rating
    # Занесение ссылки
    c4.value = link
    # Сохранение данных в таблицу Excel
    wb.save(filename)