import openpyxl as op
#переменные
name = 'Пакетик'
cost = '10P'
rating = '3,5'
url = 'https://pentagon.ru'
#имя файла
filename = 'test.xlsx'
#создание Workbook
wb = op.Workbook()
sheet = wb.active
#Вводим таблицы
c = sheet.cell(row = 1, column = 1)
c1 = sheet.cell(row = 1, column = 2)
c2 = sheet.cell(row = 1, column = 3)
c3 = sheet.cell(row = 1, column = 4)
c4 = sheet.cell(row = 1, column = 5)
#Вводим данные таблицы
c.value = "№"
c1.value = "Название"
c2.value = "Цена"
c3.value = "Рейтинг"
c4.value = "Ссылка"
#Сохраняем файл
for i in range(2, 30):
    c = sheet.cell(row=i, column=1)
    c1 = sheet.cell(row=i, column=2)
    c2 = sheet.cell(row=i, column=3)
    c3 = sheet.cell(row=i, column=4)
    c4 = sheet.cell(row=i, column=5)
    c.value = i-1
    c1.value = name
    c2.value = cost
    c3.value = rating
    c4.value = url
wb.save("test.xlsx")